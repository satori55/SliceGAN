from calendar import c
import cv2
from matplotlib.pylab import f
import matplotlib.pyplot as plt
import numpy as np
from sympy import factor, per
import tifffile as tiff
from skimage.measure import label, regionprops
from scipy.stats import wasserstein_distance
from scipy.ndimage import zoom
from torch import le
import pandas as pd
from plotnine import ggplot, aes, geom_segment, geom_point, theme_minimal, labs, theme, element_rect, element_text, element_line, geom_line, scale_shape_manual, scale_x_continuous, element_blank


def normalize_list(data):
    total = sum(data)
    if total == 0:
        return [0] * len(data)  # 如果总和为0，避免除零错误
    return [x / total for x in data]

scale_2 = False
scale_2 = True

ratio_0_7 = False
# ratio_0_7 = True


images = tiff.imread(r"D:\SliceGAN\1_stack_20_scale2_192_2.tif")
# images = tiff.imread(r"D:\SliceGAN\ctdata\fixed\4_stack_15.tif")
# images = tiff.imread(r"./gt_3.tif")
phase = np.unique(images)
# 使用scipy.ndimage.zoom进行插值，将尺寸放大n倍
# zoom_factors = (5, 5, 5)
# images = zoom(images, zoom_factors, order=0)
# print("Interpolated shape:", images.shape)

# eliminate small particles
eliminate_area = 40 # 40, 35, 10, 20

# porosity
porosity_127 = []
porosity_255 = []

for i in range(images.shape[0]):
    area_127 = np.sum(images[i, ...] == phase[1])
    area_255 = np.sum(images[i, ...] == phase[2])
    porosity_127.append(1 - (area_127 / (images[i, ...].shape[0] * images[i, ...].shape[1])))
    porosity_255.append(1 - (area_255 / (images[i, ...].shape[0] * images[i, ...].shape[1])))

average_porosity_127 = np.mean(porosity_127)
average_porosity_255 = np.mean(porosity_255)
print(f"{average_porosity_127=}")
print(f"{average_porosity_255=}")

# equivalent radius
radii = []
roundness = []
aspect_ratio = []

for i in range(images.shape[0]):
    # step2: label particles
    labeled_image, num_features = label(images[i], return_num=True)

    # step3: properties analysis
    properties = regionprops(labeled_image)

    for prop in properties:
        if prop.area <= eliminate_area:
            continue
        
        # equivalent radius roundness
        s = prop.area  # volume
        perimeter = prop.perimeter

        if perimeter == 0:
            roundness.append(0)
        else:
            roundness.append(4 * np.pi * s / perimeter ** 2)
            
        radius = (s / np.pi) ** (1/2)
        radii.append(radius)
        
        # 获取连通区域的轮廓
        coords = prop.coords
        contour = coords[:, [1, 0]].astype(np.int32)  # 转换为OpenCV格式
        
        # 获取最小旋转外接矩形
        rect = cv2.minAreaRect(contour)
        box = cv2.boxPoints(rect)
        box = box.astype(np.int32)
        
        # 计算最小旋转外接矩形的宽和高
        width_rot = np.linalg.norm(box[0] - box[1])
        height_rot = np.linalg.norm(box[1] - box[2])
        
        # 计算长宽比
        aspect_ratio_rot = width_rot / height_rot if width_rot > height_rot else height_rot / width_rot
        
        aspect_ratio.append(aspect_ratio_rot)
        # print(f"Aspect Ratio (rotated): {aspect_ratio_rot}")
    # print(f"finish{i}")
        
# equivalent to 2 times the original data
if ratio_0_7:
    radii = [r * 0.7 for r in radii]
if scale_2:
    radii = [r * 2 for r in radii]

print(f"{len(radii)=}")
print(f"{len(roundness)=}")
print(f"{radii=}")
print(f"{roundness=}")
percentiles = np.percentile(radii, [10, 50, 90])
print(f"{percentiles=}")

# mean and std
average_radius = float(np.mean(radii))
std_dev_radius = float(np.std(radii))
# print(radii)
bins = np.arange(0, 150, 2)
# bins = [0, 10, 50, 90]
hist, bin_edges = np.histogram(radii, bins=bins)

# calculate the average roundness of each particle size in the bins
roundness_bins = [[] for _ in range(len(hist))]
for r, round in zip(radii, roundness):
    for i in range(len(hist)):
        if bin_edges[i] <= r < bin_edges[i + 1]:
            roundness_bins[i].append(round)
            break

# remove the elements larger than 1 in the r of roundness_bins
for i in range(len(roundness_bins)):
    roundness_bins[i] = [r for r in roundness_bins[i] if r <= 1]
    # print(max(roundness_bins[i]))
    
# calculate the percentiles roundness
aspect_ratio_filter = []
for r, a in zip(roundness, aspect_ratio):
    if r <= 1:
        aspect_ratio_filter.append(a)

hist_aspect_ratio = np.percentile(aspect_ratio_filter, [10, 50, 90])
print(f"{hist_aspect_ratio=}")

roundness = [r for r in roundness if r <= 1]
percentiles_roundness = np.percentile(roundness, [10, 50, 90])
print(f"{percentiles_roundness=}")

# 打印每个分组的边界和计数
for i in range(len(hist)):
    print(f"分组 {i+1}: 边界 = ({bin_edges[i]}, {bin_edges[i+1]}), 计数 = {hist[i]}")

# plot
plt.figure(figsize=(10, 6))
# plt.hist(radii, bins=list(bins), color='skyblue', edgecolor='black')
plt.plot(bin_edges[:-1], hist, linestyle='-', marker='o', color='skyblue')
plt.title('Particle Radius Distribution', fontsize=24)
plt.xlabel('Radius', fontsize=20)
plt.ylabel('Frequency', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)
# plt.xlim(0, 15)
plt.grid(True)

# mean std
plt.axvline(average_radius, color='r', linestyle='dashed', linewidth=1)
plt.axvline(average_radius - std_dev_radius, color='g', linestyle='dashed', linewidth=1)
plt.axvline(average_radius + std_dev_radius, color='g', linestyle='dashed', linewidth=1)
plt.legend(['Mean Radius', 'Std Deviation'])
plt.grid(axis='x')
plt.show()

print(f"{average_radius=}")
print(f"{std_dev_radius=}")

# 绘制箱线图
plt.boxplot(radii, vert=False, patch_artist=True, 
            boxprops=dict(facecolor='bisque', color='black'),
            whiskerprops=dict(color='black'),
            capprops=dict(color='royalblue'),
            flierprops=dict(marker='o', markeredgecolor='deepskyblue', markersize=4),
            medianprops=dict(color='lightcoral'),
            showfliers=False)

plt.title('Particle Radius Distribution of Samples', fontsize=24)
plt.xlabel('Radius', fontsize=20)
plt.ylabel('Sample', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)
plt.grid(True)

plt.show()

images_2 = tiff.imread(r"./gt_1.tif")

# equivalent radius
radii2 = []

for i in range(images_2.shape[0]):
    # step2: label particles
    labeled_image, num_features = label(images_2[i], return_num=True)

    # step3: properties analysis
    properties = regionprops(labeled_image)

    for prop in properties:
        if prop.area <= 10:
            continue
        # equivalent radius
        s = prop.area  # volume
        radius = (s / np.pi) ** (1/2)
        radii2.append(radius)
        
if ratio_0_7:
    radii2 = [r * 0.7 for r in radii2]
# mean and std
average_radius2 = float(np.mean(radii2))
std_dev_radius2 = float(np.std(radii2))
# print(radii)
# bins = np.arange(0, 150, 5)
hist2, bin_edges2 = np.histogram(radii2, bins=bins)

hist = normalize_list(hist)
hist2 = normalize_list(hist2)
EMD = wasserstein_distance(hist, hist2)
print(f"{EMD=}")
print(f"{hist=}")

# plot
plt.figure(figsize=(10, 6))
# plt.hist(radii, bins=list(bins), color='skyblue', edgecolor='black')
plt.plot(bin_edges[:-1], hist, linestyle='-', marker='o', color='skyblue')
plt.plot(bin_edges[:-1], hist2, linestyle='-', marker='o', color='peru')
plt.title('Particle Radius Distribution', fontsize=24)
plt.xlabel('Radius', fontsize=20)
plt.ylabel('Frequency', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)
plt.legend(['Generated', 'Real'])
plt.grid(True)
plt.grid(axis='x')
plt.show()

print(len(hist), len(hist2))

aspect_ratio_filter = [1 / a if a != 0 else 0 for a in aspect_ratio_filter]
bins_aspect_ratio = np.arange(0, 1.1, 0.1)
hist_aspect_ratio = np.histogram(aspect_ratio_filter, bins=bins_aspect_ratio)
hist_aspect_ratio = normalize_list(hist_aspect_ratio[0])

bins_roundness = np.arange(0, 1.1, 0.1)
hist_roundness = np.histogram(roundness, bins=bins_roundness)
hist_roundness = normalize_list(hist_roundness[0])

print(len(hist_aspect_ratio), len(hist_roundness))

"""
# save to excel
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("./PSD.xlsx")
print(wb.sheetnames)
ws_1 = wb["Aspect"]
ws_2 = wb["Roundness"]
ws_3 = wb["Radius"]
column_initial = 13
row_initial = 3

for i in range(len(hist_aspect_ratio)):
    ws_1.cell(row=row_initial, column=column_initial, value=hist_aspect_ratio[i])
    ws_2.cell(row=row_initial, column=column_initial, value=hist_roundness[i])
    row_initial += 1
    
    
row_initial = 3
for i in range(len(hist)):
    ws_3.cell(row=row_initial, column=column_initial, value=hist[i])
    row_initial += 1

wb.save("./PSD.xlsx")
"""


# geom hist plot


# 创建示例数据
data = {
    'category': bin_edges[:-1],
    'start': hist,
    'end': hist2
}
df = pd.DataFrame(data)

# 创建竖直方向的杠铃图
plot = (ggplot(df, aes(y='start', yend='end', x='category', xend='category'))
        + geom_segment(size=1, color='#D3D3D3')
        + geom_point(aes(y='start'), color='#006D2C', size=2)
        + geom_point(aes(y='end'), color='#B2DF8A', size=2)
        + theme_minimal()
        + theme(panel_background=element_rect(fill='white', color='white'),
                plot_background=element_rect(fill='white', color='white'))
        + labs(title='ggplot2 geom_dumbbell with dot guide'))

# 显示图表
print(plot)


# 创建数据框架
df1 = pd.DataFrame({
    'count': hist,
    'bin_mid': 0.5 * (bin_edges[1:] + bin_edges[:-1]),
    'group': 'Generated'
})

df2 = pd.DataFrame({
    'count': hist2,
    'bin_mid': 0.5 * (bin_edges[1:] + bin_edges[:-1]),
    'group': 'Real'
})

# 合并数据框架
df = pd.concat([df1, df2])

# 创建棒棒糖图
plot = (ggplot(df, aes(x='bin_mid', y='count', color='group'))
        + geom_segment(aes(x='bin_mid', xend='bin_mid', y=0, yend='count'), size=1)
        + geom_point(size=3)
        + theme_minimal()
        + theme(panel_background=element_rect(fill='white', color='white'),
                plot_background=element_rect(fill='white', color='white'))
        + labs(title='PSD', x='Radius', y='Frequency'))

print(plot)

x_min = bin_edges.min()
x_max = bin_edges.max()

# 创建折线图
plot = (ggplot(df, aes(x='bin_mid', y='count', color='group', shape='group'))
        + geom_line(aes(group='group'), size=2)
        + geom_point(size=3)
        + theme_minimal()
        + theme(
            panel_background=element_rect(fill='white', color='white'),
            plot_background=element_rect(fill='white', color='white'),
            panel_grid_major=element_blank(),  # 去除主要网格线
            panel_grid_minor=element_blank(),  # 去除次要网格线
            panel_border=element_rect(color='black', fill=None),  # 添加四周框线
            axis_line=element_line(color='black'),  # 添加轴线
            axis_ticks_major=element_line(color='black', size=0.75),  # 添加主要刻度线
            # axis_ticks_minor=element_line(color='black', size=0.5),   # 添加次要刻度线
            axis_ticks_length_major=6,  # 主要刻度线长度
            # axis_ticks_length_minor=4,  # 次要刻度线长度
            text=element_text(size=15),
            axis_title=element_text(size=20),
            plot_title=element_text(size=24),
            legend_position=(0.95, 0.95),
            legend_justification=(1, 1),
            axis_text_x=element_text(size=22, margin={'t': 10}),  # 调整X轴文字与轴线的距离
            axis_text_y=element_text(size=22, margin={'r': 10})   # 调整Y轴文字与轴线的距离
        )
        + labs(title='Particle Radius Distribution', x='Radius', y='Frequency')
        + scale_shape_manual(values={'Generated': 'o', 'Real': 's'})  # 'o'是圆形，'s'是方形
        + scale_x_continuous(breaks=np.arange(np.floor(x_min), np.ceil(x_max) + 1, 10))  # 横轴以10为距离，并显示两端值
        + labs(color='', shape='')
)
        

plot.save("particle_radius_distribution.png", width=11, height=6, dpi=600)
