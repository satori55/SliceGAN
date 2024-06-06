import numpy as np
import tifffile as tiff
from scipy.ndimage import distance_transform_edt as distance_transform
from scipy.spatial.distance import cdist
import matplotlib.pyplot as plt
from skimage.feature import graycomatrix, graycoprops
from openpyxl import load_workbook


def two_distance_transform(image, step_size=1, max_distance=None):
    """
    Calculate the two-point correlation function of a binary image.

    Parameters:
    - image: A binary (2D) numpy array where 1s represent the feature of interest.
    - step_size: Distance step size for sampling.
    - max_distance: Maximum distance to compute the correlation for. Defaults to the half of the image diagonal.

    Returns:
    - distances: Array of distances.
    - correlation: Two-point correlation function values at the distances.
    """

    # Ensure image is binary
    image = (image > 0).astype(np.int32)

    # Calculate the distance transform
    distance = distance_transform(1 - image)

    # Maximum distance for correlation
    if max_distance is None:
        max_distance = np.hypot(*image.shape) / 2

    # Sample distances
    distances = np.arange(0, max_distance, step_size)
    correlation = []

    # Calculate correlation for each distance
    for d in distances:
        mask = np.logical_and(distance > d - step_size / 2, distance <= d + step_size / 2)
        if image[mask].size == 0:
            correlation.append(0)
        else:
            correlation.append(image[mask].size / image.size)

    return distances, correlation


def compute_two_point_correlation(image, distance):
    # 提取特征点的坐标
    features = np.argwhere(image == 1)
    
    # 如果没有足够的特征点，则直接返回0
    if len(features) < 2:
        return 0
    
    # 计算所有特征点对之间的距离
    distances = cdist(features, features, metric='euclidean')
    
    # 找到符合距离条件的点对数
    match_count = np.sum((distances >= distance) & (distances < distance + 1))
    
    # 自身与自身的距离为0，应从统计中排除，且每对点计算了两次，需要除以2
    match_count -= len(features) # 减去自身配对的数量
    match_count //= 2 # 每对点被计算了两次
    
    # 总可能的配对数量为特征点总数的组合
    total_pairs = len(features) * (len(features) - 1) // 2
    # print(match_count, total_pairs)
    
    # 计算概率
    probability = match_count / total_pairs if total_pairs > 0 else 0
    return probability


def tpcf(binary_image, distance):
    """
    计算二值图像中给定距离的二点相关函数（TPCF）。
    
    参数:
    - binary_image: 一个二值图像，用NumPy数组表示。
    - distance: 两点之间的距离。
    
    返回:
    - tpcf_value: 给定距离下的二点相关函数值。
    """
    # 初始化计数器
    match_count = 0
    total_count = 0
    
    # 获取图像的高度和宽度
    # print(binary_image.shape)
    height, width = binary_image.shape
    
    # 遍历图像中的每个点
    for y in range(height):
        for x in range(width - distance):
            # 比较当前点和给定距离后的点是否相同
            if binary_image[y, x] == binary_image[y, x + distance]:
                match_count += 1
            total_count += 1
    
    # 计算TPCF值
    tpcf_value = match_count / total_count if total_count else 0
    return tpcf_value


def calculate_glcm_features(image, distances, angles):
    """
    计算并返回给定二值图像的GLCM特征。
    
    参数:
        image (numpy.ndarray): 输入的二值图像。
        distances (list): GLCM计算中考虑的距离。
        angles (list): GLCM计算中考虑的角度。
        
    返回:
        features (dict): 包含对比度、同质性、能量和相关性特征的字典。
    """
    # 计算灰度共生矩阵
    glcm = graycomatrix(image, distances=distances, angles=angles, levels=2, symmetric=True, normed=True)
    # temp = graycoprops(glcm, 'contrast')
    # print(temp)
    
    # 提取纹理特征
    features = {
        'contrast': graycoprops(glcm, 'contrast'),
        'homogeneity': graycoprops(glcm, 'homogeneity'),
        'energy': graycoprops(glcm, 'energy'),
        'correlation': graycoprops(glcm, 'correlation')
    }
    
    return features


# Example usage:
if __name__ == "__main__":
    wb = load_workbook("C:/Users/SATORI/My drive/GNN/evaluation.xlsx")
    print(wb.sheetnames)
    ws = wb["Sheet1"]
    # Create a random binary image
    # np.random.seed(0)
    # image = np.random.choice([0, 1], size=(100, 100), p=[0.7, 0.3])
    images: np.ndarray = tiff.imread('5slices_RMS_3.tif')
    row_index = 18
    contrasts = []
    homogeneities = []
    energies = []
    correlations = []
    for i in range(images.shape[0]):
        image = images[i, ...] / 255
        # print(image)
        # distances, correlation = two_distance_transform(image, max_distance=10)

        # # Plot the two-point correlation function
        # plt.figure(figsize=(8, 5))
        # plt.plot(distances, correlation)
        # plt.xlabel('Distance')
        # plt.ylabel('Correlation')
        # plt.title('Two-Point Correlation Function')
        # plt.show()
        

        # 计算距离为1时的两点相关函数
        # image = np.array([
        # [1, 0, 1, 0, 1],
        # [0, 1, 0, 1, 0],
        # [1, 0, 1, 0, 1],
        # [0, 1, 0, 1, 0],
        # [1, 0, 1, 0, 1]
        # ])
        distance = 5
        probability = compute_two_point_correlation(image, distance)
        tcpf_res = tpcf(image, distance)
        glcm_features = calculate_glcm_features(image.astype(np.int32), distances=[1], angles=[0, np.pi/4, np.pi/2, 3*np.pi/4])
        # print(glcm_features)
        # print(f"在距离{distance}下两点同时为1的概率是{tcpf_res}")
        # print(f"在距离{distance}下两点同时为1的概率是{probability}")
        #计算所有slice的GLCM特征平均值
        contrasts.append(glcm_features['contrast'])
        homogeneities.append(glcm_features['homogeneity'])
        energies.append(glcm_features['energy'])
        correlations.append(glcm_features['correlation'])
        
    mean_contrasts = np.mean(contrasts, axis=0)
    mean_homogeneities = np.mean(homogeneities, axis=0)
    mean_energies = np.mean(energies, axis=0)
    mean_correlations = np.mean(correlations, axis=0)
    # print('contrast:', np.mean(contrasts, axis=0))
    # print('homogeneity:', np.mean(homogeneities))
    # print('energy:', np.mean(energies))
    # print('correlation:', np.mean(correlations))

    for i in range(4):
        ws.cell(row=row_index, column=5 + 4 * i).value = mean_contrasts[0][i]
        ws.cell(row=row_index, column=6 + 4 * i).value = mean_homogeneities[0][i]
        ws.cell(row=row_index, column=7 + 4 * i).value = mean_energies[0][i]
        ws.cell(row=row_index, column=8 + 4 * i).value = mean_correlations[0][i]
    wb.save("C:/Users/SATORI/My drive/GNN/evaluation.xlsx")
    